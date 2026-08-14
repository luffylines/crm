import io
import json
import os
from datetime import datetime

import pandas as pd

import app as app_module


def test_contactout_search_route_accepts_manual_search_payload():
    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    response = client.post(
        "/contactout/search",
        json={
            "first_name": "Ava",
            "last_name": "Ruiz",
            "company": "Sumware Software",
            "title": "Manager",
            "include_contacts": False,
        },
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert "ok" in payload
    assert "found" in payload


def test_contactout_falls_back_to_apollo_when_demo_profile_is_returned(monkeypatch):
    app_module.CONTACTOUT_API_KEY = "demo-key"
    app_module.APOLLO_API_KEY = "apollo-key"

    class DummyResponse:
        def __init__(self, status_code, payload):
            self.status_code = status_code
            self._payload = payload

        def json(self):
            return self._payload

    def fake_post(url, **kwargs):
        if url.startswith("https://api.contactout.com"):
            return DummyResponse(200, {
                "profile": {
                    "full_name": "Example Person",
                    "headline": "Manager, Business Operations & Marketing at OBM",
                    "company": {"name": "Legros, Smitham and Kessler", "website": "www.legros.com"},
                    "url": "https://www.linkedin.com/in/example-person",
                }
            })
        if url.startswith("https://api.apollo.io"):
            return DummyResponse(200, {
                "people": [{
                    "name": "Brandon Callor",
                    "title": "MSW",
                    "organization_name": "Acadain Counseling, Inc",
                    "email": "brandon@acadain.org",
                    "linkedin_url": "https://www.linkedin.com/in/brandon-callow",
                }]
            })
        raise AssertionError(f"Unexpected URL: {url}")

    monkeypatch.setattr(app_module.req, "post", fake_post)

    result = app_module.contactout_enrich_person({
        "First Name": "Brandon",
        "Last Name": "Callor",
        "Company": "Acadain Counseling, Inc",
        "Title": "MSW",
    }, include_contacts=False)

    assert result["ok"] is True
    assert result["found"] is True
    assert result["profile"]["name"] == "Brandon Callor"
    assert result["provider"] == "apollo"


def test_missing_contactout_key_returns_clean_message():
    original_co = app_module.CONTACTOUT_API_KEY
    original_apollo = app_module.APOLLO_API_KEY
    app_module.CONTACTOUT_API_KEY = ""
    app_module.APOLLO_API_KEY = ""
    try:
        result = app_module.contactout_enrich_person({
            "First Name": "James",
            "Last Name": "Miller",
        }, include_contacts=False)
        assert result["ok"] is True
        assert result["found"] is False
        assert "Add a real CONTACTOUT_API_KEY" in result["message"]
    finally:
        app_module.CONTACTOUT_API_KEY = original_co
        app_module.APOLLO_API_KEY = original_apollo


def test_validated_date_normalizes_to_iso_string():
    assert app_module.normalize_validated_date("2024-05-23") == "2024-05-23"
    assert app_module.normalize_validated_date("08/12/2026") == "2026-08-12"
    assert app_module.normalize_validated_date("") == datetime.now().strftime("%Y-%m-%d")


def test_upload_maps_company_name_column_to_company_field(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    df = pd.DataFrame({
        "Company Name": ["Beacon Capital Management"],
        "Website": ["https://example.com"],
        "Lead Ranking": ["better"],
    })
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    response = client.post(
        "/upload",
        data={"file": (buf, "company_name.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    with client.session_transaction() as sess:
        key = sess.get("file_key")

    store = app_module.stores.get(f"{sess['username']}::{key}")
    assert store is not None
    assert store["df"]["Company"].tolist() == ["Beacon Capital Management"]


def test_upload_keeps_multiple_files_with_same_name(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    def post_file(name, company):
        df = pd.DataFrame({"Company": [company], "Lead Ranking": ["good"]})
        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        return client.post(
            "/upload",
            data={"file": (buf, name)},
            content_type="multipart/form-data",
        )

    first = post_file("repeat.xlsx", "Acme")
    second = post_file("repeat.xlsx", "Beta")

    assert first.status_code == 200
    assert second.status_code == 200

    files = client.get("/files").get_json()["files"]
    assert len(files) >= 2
    assert any(f["filename"] == "repeat.xlsx" for f in files)


def test_upload_discovers_arbitrary_sheet_names_without_hardcoded_assumptions(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    workbook_path = os.path.join(tmp_path, "dynamic_workbook.xlsx")
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame({"Company": ["Alpha"], "Lead Ranking": ["good"]}).to_excel(writer, sheet_name="Raw Data", index=False)
        pd.DataFrame({"Company": ["Bravo"], "Lead Ranking": ["best"]}).to_excel(writer, sheet_name="Validation", index=False)
        pd.DataFrame({"Company": ["Charlie"], "Lead Ranking": ["better"]}).to_excel(writer, sheet_name="Completed Leads", index=False)

    with open(workbook_path, "rb") as fh:
        response = client.post(
            "/upload",
            data={"file": (io.BytesIO(fh.read()), "dynamic_workbook.xlsx")},
            content_type="multipart/form-data",
        )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["total"] == 1

    with client.session_transaction() as sess:
        key = sess.get("file_key")

    open_response = client.get(f"/open/{key}")
    assert open_response.status_code == 200
    open_payload = open_response.get_json()
    sheet_names = [ws["name"] for ws in open_payload["worksheets"]]
    assert sheet_names == ["Raw Data", "Validation", "Completed Leads"]


def test_save_row_debounces_excel_disk_writes(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"
        sess["file_key"] = "demo"

    draft_path = os.path.join(tmp_path, "uploader_demo_draft.xlsx")
    pd.DataFrame({"Company": ["Acme"], "Lead Ranking": ["good"], "Website": ["https://example.com"]}).to_excel(draft_path, index=False)

    store = {
        "df": pd.DataFrame({"Company": ["Acme"], "Lead Ranking": ["good"], "Website": ["https://example.com"]}),
        "original_df": pd.DataFrame({"Company": ["Acme"], "Lead Ranking": ["good"], "Website": ["https://example.com"]}),
        "validated": set(),
        "filename": "demo.xlsx",
    }
    app_module.stores["uploader::demo"] = store

    calls = {"count": 0}

    def fake_write(df, path):
        calls["count"] += 1
        return True

    monkeypatch.setattr(app_module, "safe_write_excel", fake_write)

    first = client.post("/save/0", json={"Company": "Acme", "Lead Ranking": "best", "Website": "https://example.com"})
    second = client.post("/save/0", json={"Company": "Acme", "Lead Ranking": "best", "Website": "https://example.com"})

    assert first.status_code == 200
    assert second.status_code == 200
    assert calls["count"] <= 1


def test_open_file_lists_worksheets_before_user_picks_one(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    master = pd.DataFrame({"Company": ["Alpha"], "Lead Ranking": ["best"]})
    reval = pd.DataFrame({"Company": ["Bravo"], "Lead Ranking": ["good"]})

    draft_path = os.path.join(tmp_path, "uploader_selection_draft.xlsx")
    with pd.ExcelWriter(draft_path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Masterfile", index=False)
        reval.to_excel(writer, sheet_name="Christian For Reval", index=False)

    response = client.get("/open/selection")
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["worksheets"][0]["name"] == "Masterfile"
    assert any(ws["name"] == "Christian For Reval" for ws in payload["worksheets"])


def test_list_files_uses_cached_summary_without_reopening_every_workbook(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    draft_name = "uploader_cached_draft.xlsx"
    draft_path = os.path.join(tmp_path, draft_name)
    pd.DataFrame({"Company": ["Acme"], "_validated": ["1"]}).to_excel(draft_path, index=False)

    summary_path = os.path.splitext(draft_path)[0] + ".summary.json"
    with open(summary_path, "w", encoding="utf-8") as summary_file:
        json.dump({"total": 1, "done": 1, "status": "Completed", "modified_ts": os.path.getmtime(draft_path)}, summary_file)

    import openpyxl
    calls = {"count": 0}
    original_load_workbook = openpyxl.load_workbook

    def fake_load_workbook(*args, **kwargs):
        calls["count"] += 1
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load_workbook)

    response = client.get("/files")
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["files"][0]["status"] == "Completed"
    assert calls["count"] == 0


def test_open_single_worksheet_skips_selection_modal(monkeypatch, tmp_path):
    """When workbook has exactly 1 worksheet, skip modal and open directly."""
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    workbook_path = os.path.join(tmp_path, "single_sheet.xlsx")
    pd.DataFrame({"Company": ["Alpha"], "Lead Ranking": ["good"]}).to_excel(
        workbook_path, sheet_name="Sheet1", index=False
    )

    with open(workbook_path, "rb") as fh:
        response = client.post(
            "/upload",
            data={"file": (io.BytesIO(fh.read()), "single_sheet.xlsx")},
            content_type="multipart/form-data",
        )

    assert response.status_code == 200

    with client.session_transaction() as sess:
        key = sess.get("file_key")

    open_response = client.get(f"/open/{key}")
    assert open_response.status_code == 200
    payload = open_response.get_json()
    
    # Should NOT have worksheets array; should have total, resume_index, sheet_name
    assert "worksheets" not in payload
    assert payload["total"] == 1
    assert payload["sheet_name"] == "Sheet1"
    assert payload["resumed"] == True


def test_open_multiple_worksheets_shows_selection_modal(monkeypatch, tmp_path):
    """When workbook has 2+ worksheets, show modal with worksheet list."""
    monkeypatch.setattr(app_module, "DRAFT_DIR", str(tmp_path))
    os.makedirs(tmp_path, exist_ok=True)

    client = app_module.app.test_client()
    with client.session_transaction() as sess:
        sess["username"] = "uploader"

    workbook_path = os.path.join(tmp_path, "multi_sheet.xlsx")
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame({"Company": ["Alpha"], "Lead Ranking": ["good"]}).to_excel(
            writer, sheet_name="Masterfile", index=False
        )
        pd.DataFrame({"Company": ["Bravo"], "Lead Ranking": ["best"]}).to_excel(
            writer, sheet_name="Christian For Reval", index=False
        )

    with open(workbook_path, "rb") as fh:
        response = client.post(
            "/upload",
            data={"file": (io.BytesIO(fh.read()), "multi_sheet.xlsx")},
            content_type="multipart/form-data",
        )

    assert response.status_code == 200

    with client.session_transaction() as sess:
        key = sess.get("file_key")

    open_response = client.get(f"/open/{key}")
    assert open_response.status_code == 200
    payload = open_response.get_json()
    
    # Should have worksheets array with all sheet names
    assert "worksheets" in payload
    assert len(payload["worksheets"]) == 2
    sheet_names = [ws["name"] for ws in payload["worksheets"]]
    assert "Masterfile" in sheet_names
    assert "Christian For Reval" in sheet_names

