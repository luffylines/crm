from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
import pandas as pd
import requests as req
import re
import io
import os
import json
import pickle
import stat
import threading
from datetime import datetime, timedelta, timezone
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import time
from functools import wraps

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - fallback when python-dotenv is not installed
    def load_dotenv(*args, **kwargs):
        return False

dotenv_path = os.path.join(os.path.dirname(__file__), ".env")
load_dotenv(dotenv_path)

# Import models and auth helpers
from models import db, User, ActivityLog, UploadedFile
from auth_helpers import (
    login_required as auth_login_required,
    admin_required,
    viewer_required,
    get_current_user,
    log_activity,
    get_client_ip
)

def _clean_api_key(value):
    if value is None:
        return ""
    clean = str(value).strip()
    if not clean:
        return ""
    lowered = clean.lower()
    placeholder_markers = [
        "paste_", "your_", "demo", "example", "placeholder", "replace_me", "changeme"
    ]
    if any(marker in lowered for marker in placeholder_markers):
        return ""
    return clean


CONTACTOUT_API_KEY = _clean_api_key(os.getenv("CONTACTOUT_API_KEY"))
CONTACTOUT_BASE_URL = "https://api.contactout.com"
APOLLO_API_KEY = _clean_api_key(os.getenv("APOLLO_API_KEY"))
APOLLO_BASE_URL = "https://api.apollo.io"
PDL_API_KEY = _clean_api_key(os.getenv("PDL_API_KEY"))
PDL_BASE_URL = "https://api.peopledatalabs.com/v5"

app = Flask(__name__)

app.config["SECRET_KEY"] = os.getenv("FLASK_SECRET_KEY") or "dev-secret-key-change-me"

# Database configuration
DATABASE_URL = os.getenv("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    # Fix postgres:// to postgresql:// for modern SQLAlchemy
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL or "sqlite:///crm.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Initialize database
db.init_app(app)


def ensure_default_users(users_file_path=None):
    """Seed the default login accounts from users.json when the DB is empty."""
    default_path = users_file_path or os.path.join(os.path.dirname(__file__), "users.json")
    if not os.path.exists(default_path):
        return False

    if User.query.count() > 0:
        return False

    try:
        with open(default_path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (OSError, ValueError):
        return False

    created = 0
    for user_data in data.get("users", []) or []:
        username = str(user_data.get("username", "")).strip().lower()
        password = user_data.get("password", "")

        if not username or not password:
            continue

        if User.query.filter_by(username=username).first():
            continue

        role = "Admin" if username == "admin" else "User"
        user = User(username=username, role=role, is_active=True)
        user.set_password(password)
        db.session.add(user)
        created += 1

    if created:
        db.session.commit()

    return created > 0


# Create tables on app startup
with app.app_context():
    db.create_all()
    ensure_default_users()

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=86400
)

DRAFT_DIR = "drafts"
os.makedirs(DRAFT_DIR, exist_ok=True)

# Per-file in-memory store: { session_key: { df, original_df, filename, validated } }
stores = {}


def normalize_company_fields(df):
    if df is None:
        return df

    if "Company" not in df.columns:
        df["Company"] = ""

    company_candidates = [
        "Company Name", "company_name", "Company name", "company name",
        "Organization Name", "organization_name", "CompanyName"
    ]

    for candidate in company_candidates:
        if candidate not in df.columns:
            continue
        if df["Company"].isna().all() or df["Company"].astype(str).str.strip().eq("").all():
            df["Company"] = df[candidate].fillna("")
            continue
        mask = df["Company"].astype(str).str.strip().eq("")
        if mask.any():
            df.loc[mask, "Company"] = df.loc[mask, candidate].fillna("")

    return df


def canonicalize_excel_columns(df):
    if df is None:
        return df

    rename_map = {}
    field_aliases = {
        "Record Id": ["Record Id", "RecordID", "record_id", "Record ID"],
        "About Company": ["About Company", "About / Company Description", "About", "Company Description"],
        "No. of Employees": ["No. of Employees", "No of Employees", "Employees", "Employee Count", "Employee count"],
        "Company Industry": ["Company Industry", "Industry", "company industry"],
    }

    for canonical, aliases in field_aliases.items():
        candidates = [alias for alias in aliases if alias in df.columns]
        if not candidates:
            continue
        primary = canonical
        if primary not in df.columns:
            rename_map[candidates[0]] = primary
        for alias in candidates:
            if alias == primary:
                continue
            if primary in df.columns:
                if alias in df.columns and primary in df.columns:
                    df[primary] = df[primary].fillna("")
                    df.loc[df[primary].astype(str).str.strip().eq(""), primary] = df.loc[df[primary].astype(str).str.strip().eq(""), alias].fillna("")
                df.drop(columns=[alias], inplace=True, errors='ignore')
            else:
                rename_map[alias] = primary

    if rename_map:
        df.rename(columns=rename_map, inplace=True)

    for col in ["About Company", "No. of Employees", "Company Industry"]:
        if col not in df.columns:
            df[col] = ""

    return df


def login_required(f):
    """Wrapper for backward compatibility - uses new auth system."""
    return auth_login_required(f)


def get_username():
    return session.get("username", "")

ACCEPTED_TITLES = [
    "office manager", "it manager", "cfo", "coo", "controller",
    "general manager", "administrative assistant", "operations manager",
    "chief financial officer", "chief operating officer"
]

geolocator = Nominatim(user_agent="revalidation_tool")


def normalize_lead_ranking_value(value):
    return str(value).strip().lower().replace("_", " ")


def is_completed_lead_ranking(value):
    return normalize_lead_ranking_value(value) in {"bad", "good", "better", "best", "manual review"}


def is_row_validated(row):
    """
    Determine if a row is already validated.
    A row is considered validated if it has a Lead Ranking value.
    """
    return is_completed_lead_ranking(row.get("Lead Ranking", ""))


def normalize_workbook_sheet(df):
    if df is None:
        return df
    cleaned = df.fillna("")
    cleaned = canonicalize_excel_columns(cleaned)
    if "be" in cleaned.columns:
        cleaned.rename(columns={"be": "Lead Ranking"}, inplace=True)
    cleaned = normalize_company_fields(cleaned)
    if "Lead Ranking" not in cleaned.columns:
        cleaned["Lead Ranking"] = ""
    if "Validated Date" not in cleaned.columns:
        cleaned["Validated Date"] = ""
    return cleaned


def get_workbook_sheet_map(file_like):
    try:
        file_like.seek(0)
        sheets = pd.read_excel(file_like, sheet_name=None, dtype=str)
    except Exception:
        try:
            file_like.seek(0)
            single = pd.read_excel(file_like, dtype=str).fillna("")
            sheets = {"Sheet1": single}
        except Exception:
            return {}

    normalized = {}
    for name, sheet_df in sheets.items():
        normalized[name] = normalize_workbook_sheet(sheet_df)
    return normalized


def detect_working_sheet_name(sheet_map):
    if not sheet_map:
        return None

    for name, df in sheet_map.items():
        if df is None:
            continue
        normalized = normalize_workbook_sheet(df.copy())
        if "Lead Ranking" not in normalized.columns:
            continue
        values = normalized["Lead Ranking"].astype(str).str.strip().str.lower()
        if values.isin(["bad", "good", "better", "best"]).any():
            return name

    for name in sheet_map:
        if name:
            return name
    return None


def build_filtered_queue(df, selected_validator):
    """
    Build a filtered queue of row indices based on the selected validator.

    This keeps completed rows visible in the Progress list while still tracking
    which ones are already validated. The validation flow resumes on the first
    incomplete row, but the queue includes all rows assigned to the active
    validator so their saved Lead Ranking values remain visible and clickable.
    """
    validated = set()
    filtered_queue = []

    if "Validated By" not in df.columns:
        df["Validated By"] = ""
    if "Lead Ranking" not in df.columns:
        df["Lead Ranking"] = ""

    for idx in range(len(df)):
        row = df.iloc[idx]
        lead_ranking = str(row.get("Lead Ranking", "")).strip()
        validated_by = str(row.get("Validated By", "")).strip()

        if lead_ranking and is_completed_lead_ranking(lead_ranking):
            if not selected_validator or validated_by == selected_validator:
                validated.add(idx)

        if selected_validator:
            if validated_by == selected_validator:
                filtered_queue.append(idx)
        else:
            filtered_queue.append(idx)

    return validated, filtered_queue


def build_worksheet_selection_payload(key, sheet_map):
    worksheets = []
    for name, df in sheet_map.items():
        if df is None:
            continue
        validated_count = 0
        validators = set()
        if "Lead Ranking" in df.columns:
            validated_count = int(df["Lead Ranking"].astype(str).str.strip().str.lower().isin(["bad", "good", "better", "best"]).sum())
        if "Validated By" in df.columns:
            for val in df["Validated By"]:
                val_str = str(val).strip() if val else ""
                if val_str:
                    validators.add(val_str)
        worksheets.append({
            "name": name,
            "rows": len(df),
            "validated_count": validated_count,
            "validators": sorted(list(validators)),
            "validated_by_counts": {v: int((df["Validated By"].astype(str) == v).sum()) for v in validators},
            "description": "Worksheet data"
        })
    return {"key": key, "worksheets": worksheets}


def safe_write_excel(df, draft_path, masterfile_df=None, working_sheet_name="Sheet1", workbook_sheets=None):
    """Write Excel file while preserving all workbook sheets when provided."""
    os.makedirs(os.path.dirname(draft_path) or ".", exist_ok=True)

    if os.path.exists(draft_path):
        try:
            os.chmod(draft_path, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IWGRP | stat.S_IROTH | stat.S_IWOTH)
        except Exception:
            pass

    base, ext = os.path.splitext(draft_path)
    ext = ext or '.xlsx'
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    tmp_path = f"{base}.tmp_{stamp}{ext}"
    fallback_path = f"{base}.autosave_{stamp}{ext}"

    def write_workbook(target_path):
        if workbook_sheets is not None:
            with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
                for sheet_name, sheet_df in workbook_sheets.items():
                    if sheet_df is None:
                        continue
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            return

        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            if masterfile_df is not None:
                masterfile_df.to_excel(writer, sheet_name="Masterfile", index=False)
            df.to_excel(writer, sheet_name=working_sheet_name, index=False)

    try:
        write_workbook(tmp_path)
        try:
            os.replace(tmp_path, draft_path)
        except PermissionError:
            try:
                write_workbook(fallback_path)
            except Exception:
                pass
            return False
        except OSError:
            try:
                write_workbook(fallback_path)
            except Exception:
                pass
            return False
        return True
    except Exception:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        try:
            write_workbook(fallback_path)
        except Exception:
            pass
        return False
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def is_safe_key(key):
    """
    Validate that a file key does not contain path traversal sequences.
    Prevents attacks like ../user1_file, ..\\user1_file, etc.
    """
    if not key:
        return False
    safe_key = os.path.basename(str(key))
    traversal_patterns = ['..', '~', '//', '\\\\']
    for pattern in traversal_patterns:
        if pattern in safe_key:
            return False
    return safe_key == str(key)


def build_user_draft_path(username, key):
    """Build a deterministic draft path for a specific user without depending on request session state."""
    user = str(username or "").strip()
    if not user or not key:
        return None
    if not is_safe_key(key):
        return None
    safe_key = os.path.basename(str(key))
    return os.path.join(DRAFT_DIR, f"{user}_{safe_key}_draft.xlsx")


def get_user_draft_path(key):
    """
    Generate a deterministic file path for a user's draft file.
    Uses the file_key directly (no auto-incrementing).
    This should be used for opening/saving/deleting existing files.
    
    Format: {DRAFT_DIR}/{username}_{key}_draft.xlsx
    """
    return build_user_draft_path(get_username(), key)


def get_draft_path(filename):
    """Generate a new draft path (with auto-increment if exists).
    
    WARNING: Only use this for NEW file uploads.
    For existing files, use get_user_draft_path(key) instead.
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    user = get_username()
    prefix = f"{user}_" if user else ""
    draft_name = f"{prefix}{base}_draft.xlsx"
    draft_path = os.path.join(DRAFT_DIR, draft_name)
    if not os.path.exists(draft_path):
        return draft_path

    counter = 2
    while True:
        alt_name = f"{prefix}{base}_{counter}_draft.xlsx"
        alt_path = os.path.join(DRAFT_DIR, alt_name)
        if not os.path.exists(alt_path):
            return alt_path
        counter += 1


def get_draft_summary_path(draft_path):
    return os.path.splitext(draft_path)[0] + ".summary.json"


def build_draft_summary(df, validated=None):
    validated = validated or set()
    total_rows = len(df) if df is not None else 0
    done = sum(1 for i in validated if i < total_rows)
    status = "Completed" if total_rows > 0 and done >= total_rows else "In Progress"
    return {
        "total": total_rows,
        "done": done,
        "status": status,
        "modified_ts": time.time(),
    }


def write_draft_summary(draft_path, df=None, validated=None):
    if not draft_path:
        return

    try:
        summary = build_draft_summary(df, validated)
        summary_path = get_draft_summary_path(draft_path)
        with open(summary_path, "w", encoding="utf-8") as summary_file:
            json.dump(summary, summary_file)
    except Exception:
        pass


def read_draft_summary(draft_path):
    summary_path = get_draft_summary_path(draft_path)
    if not os.path.exists(summary_path):
        return None

    try:
        with open(summary_path, "r", encoding="utf-8") as summary_file:
            return json.load(summary_file)
    except Exception:
        return None


def get_session_key():
    if "file_key" not in session:
        session["file_key"] = None
    user = get_username()
    key = session.get("file_key")
    if key and user:
        return f"{user}::{key}"
    return key


def get_store():
    key = get_session_key()  # format: "username::filekey"
    if not key:
        return None
    if key in stores:
        return stores[key]
    # Try loading from pickle
    parts = key.split("::", 1)
    if len(parts) == 2:
        pkl_name = f"{parts[0]}_{parts[1]}.pkl"
    else:
        pkl_name = f"{key}.pkl"
    pkl = os.path.join(DRAFT_DIR, pkl_name)
    if os.path.exists(pkl):
        try:
            with open(pkl, "rb") as f:
                stores[key] = pickle.load(f)
            return stores[key]
        except Exception:
            pass
    return None


def save_store(key, store):
    stores[key] = store
    # key is "username::filekey" — derive pickle path
    parts = key.split("::", 1)
    if len(parts) == 2:
        pkl_name = f"{parts[0]}_{parts[1]}.pkl"
    else:
        pkl_name = f"{key}.pkl"
    pkl = os.path.join(DRAFT_DIR, pkl_name)
    try:
        with open(pkl, "wb") as f:
            pickle.dump(store, f)
    except Exception:
        pass


def flush_store_to_disk(store, *, file_key=None, force=False):
    if not file_key:
        return False

    draft_path = get_user_draft_path(file_key)
    if not draft_path:
        return False

    now = time.time()
    last_write = session.get("last_disk_write_ts")
    if not force and last_write is not None and (now - float(last_write)) < 2.0:
        save_store(get_session_key(), store)
        return False

    working_sheet_name = store.get("working_sheet_name") or detect_working_sheet_name(store.get("sheet_map")) or "Sheet1"
    sheet_map = store.get("sheet_map")
    if sheet_map:
        for name, df in list(sheet_map.items()):
            if df is not None and name == working_sheet_name:
                sheet_map[name] = store["df"].copy()
    else:
        sheet_map = {}
        sheet_map[working_sheet_name] = store["df"].copy()

    draft_df = store["df"].copy()
    draft_df["_validated"] = draft_df.index.map(lambda i: "1" if i in store["validated"] else "")
    if sheet_map and working_sheet_name in sheet_map:
        sheet_map[working_sheet_name] = draft_df

    try:
        safe_write_excel(
            draft_df,
            draft_path,
            masterfile_df=store.get("masterfile_df"),
            working_sheet_name=working_sheet_name,
            workbook_sheets=sheet_map,
        )
    except TypeError as exc:
        if "unexpected keyword argument" in str(exc):
            safe_write_excel(draft_df, draft_path)
        else:
            raise

    write_draft_summary(draft_path, store["df"], store["validated"])
    session["last_disk_write_ts"] = now
    save_store(get_session_key(), store)
    return True


def clean_phone(phone):
    if not phone:
        return ""
    digits = re.sub(r'\D', '', str(phone))
    if digits.startswith('1') and len(digits) == 11:
        digits = digits[1:]
    return digits


TOLL_FREE = {800, 833, 844, 855, 866, 877, 888}


def validate_phone(phone):
    digits = clean_phone(phone)
    if len(digits) != 10:
        return False, "Phone must be exactly 10 digits"
    prefix = int(digits[:3])
    if prefix in TOLL_FREE:
        return False, f"Toll-free numbers ({prefix}) not accepted"
    return True, "Valid"


def validate_email(email):
    if not email or str(email).strip() == "" or str(email).strip().lower() == "nan":
        return False, "Missing"
    if re.match(r'^[\w\.-]+@[\w\.-]+\.\w{2,}$', str(email).strip()):
        return True, "Valid"
    return False, "Invalid format"


def normalize_validated_date(value):
    if value is None:
        return datetime.now().strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return datetime.now().strftime("%Y-%m-%d")
    try:
        parsed = datetime.strptime(text, "%Y-%m-%d")
        return parsed.strftime("%Y-%m-%d")
    except ValueError:
        pass

    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return datetime.now().strftime("%Y-%m-%d")


def validate_website(url):
    if not url or str(url).strip() == "" or str(url).strip().lower() == "nan":
        return False, "Missing"
    try:
        url = str(url).strip()
        if not url.startswith("http"):
            url = "https://" + url
        resp = req.get(url, timeout=7, allow_redirects=True)
        # Accept 2xx and 3xx as valid; allow 403/404 (site may block automation but exists)
        if resp.status_code < 400:
            return True, "Active"
        elif resp.status_code in [403, 404]:
            return True, f"HTTP {resp.status_code} (restricted access)"
        return False, f"HTTP {resp.status_code}"
    except Exception:
        return False, "Unreachable"


def validate_address(street, city, state, zipcode):
    if not all([street, city, state, zipcode]):
        return False, "Incomplete address"
    addr = str(street).strip().upper()
    if re.search(r'\bP\.?\s*O\.?\s*BOX\b', addr):
        return False, "PO BOX not accepted"
    full_address = f"{street}, {city}, {state} {zipcode}, USA"
    try:
        time.sleep(1)
        location = geolocator.geocode(full_address, timeout=10)
        if location:
            return True, "Valid"
        return False, "Address not found"
    except GeocoderTimedOut:
        return False, "Geocoder timeout"
    except Exception as e:
        return False, f"Error: {str(e)}"


def validate_title(title):
    if not title or str(title).strip().lower() == "nan":
        return False, "Missing"
    title_lower = str(title).strip().lower()
    for accepted in ACCEPTED_TITLES:
        if accepted in title_lower:
            return True, "Valid"
    return False, "Title not in accepted list"


def calculate_rank(row, validations):
    """Calculate the final Lead Ranking using validated contact and enriched company details."""
    def has_value(*keys):
        for key in keys:
            value = str(row.get(key, "")).strip()
            if value and value.lower() != "nan":
                return True
        return False

    company_ok = has_value("Company", "Company Name")
    first_name_ok = has_value("First Name", "first_name")
    last_name_ok = has_value("Last Name", "last_name")
    phone_ok = has_value("Phone")
    email_ok = has_value("Email")
    website_ok = validations.get("website", {}).get("valid", False)
    about_ok = has_value("About Company", "About / Company Description", "About", "Company Description")
    industry_ok = has_value("Company Industry", "Industry")
    employees_ok = has_value("No. of Employees", "No of Employees", "Employee Count", "Employees")
    alt_contact = has_value("Alt. Contact Info", "Alternate Contact Info")
    alt_phone = has_value("Alternate Phone", "Alt. Phone")
    company_evidence = about_ok or industry_ok or employees_ok

    if not company_ok:
        return "Bad", "Company Name is missing"
    if not (first_name_ok or last_name_ok):
        return "Bad", "First Name and Last Name are both missing"
    if not (phone_ok or email_ok):
        return "Bad", "Phone and Email are both missing"
    if not website_ok and not company_evidence:
        return "Bad", "Website is missing and there is no useful company information"

    complete_contact = company_ok and first_name_ok and last_name_ok and phone_ok and email_ok
    if complete_contact and website_ok and about_ok and industry_ok and employees_ok and alt_contact and alt_phone:
        return "Best", "Complete contact and company information available"

    if complete_contact and website_ok and about_ok and industry_ok and employees_ok:
        if not alt_contact or not alt_phone:
            return "Better", "Main contact and company information are usable but alternate contact details are missing"

    if complete_contact and not website_ok and company_evidence:
        return "Good", "Useful company information was identified even though Website is missing"

    if complete_contact and not (website_ok or company_evidence):
        return "Manual Review", "Important company info could not be verified reliably"

    if not company_evidence and not (alt_contact or alt_phone):
        return "Manual Review", "Lead requires manual review before final ranking can be confirmed"

    return "Better", "Main contact and company information are usable but not fully complete"


def ensure_company_enrichment_columns(df):
    df = canonicalize_excel_columns(df)
    for col in ["About Company", "Company Industry", "No. of Employees"]:
        if col not in df.columns:
            df[col] = ""
    if "Manual Review" not in df.columns:
        df["Manual Review"] = ""
    return df


def apply_lead_ranking_conditional_formatting(workbook_path):
    try:
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill

        workbook = load_workbook(workbook_path)
        for worksheet in workbook.worksheets:
            if worksheet.max_row < 2:
                continue
            if "Lead Ranking" not in worksheet[1].values:
                continue

            lead_col_index = None
            for idx, cell in enumerate(worksheet[1], start=1):
                if str(cell.value).strip() == "Lead Ranking":
                    lead_col_index = idx
                    break
            if lead_col_index is None:
                continue

            last_col = worksheet.max_column
            last_row = worksheet.max_row
            for row_idx in range(2, last_row + 1):
                cell = worksheet.cell(row=row_idx, column=lead_col_index)
                value = str(cell.value).strip() if cell.value is not None else ""
                if value.lower() == "best":
                    fill = PatternFill(fill_type="solid", fgColor="006100")
                elif value.lower() == "better":
                    fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
                elif value.lower() == "good":
                    fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
                elif value.lower() == "bad":
                    fill = PatternFill(fill_type="solid", fgColor="FFC000")
                elif value.lower() == "manual review":
                    fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                else:
                    fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                for col_idx in range(1, last_col + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fill

        workbook.save(workbook_path)
    except Exception:
        pass


def run_validations(row):
    phone_valid, phone_msg   = validate_phone(row.get("Phone", ""))
    email_valid, email_msg   = validate_email(row.get("Email", ""))
    website_valid, website_msg = validate_website(row.get("Website", ""))
    address_valid, address_msg = validate_address(
        row.get("Street", ""), row.get("City", ""),
        row.get("State", ""), row.get("Zip Code", "")
    )
    title_valid, title_msg   = validate_title(row.get("Title", ""))
    validations = {
        "phone":   {"valid": phone_valid,   "msg": phone_msg},
        "email":   {"valid": email_valid,   "msg": email_msg},
        "website": {"valid": website_valid, "msg": website_msg},
        "address": {"valid": address_valid, "msg": address_msg},
        "title":   {"valid": title_valid,   "msg": title_msg},
    }
    suggested_rank, rank_reason = calculate_rank(row, validations)
    return validations, suggested_rank, rank_reason


def _company_field(row, *keys):
    for key in keys:
        value = row.get(key, "")
        if value is None:
            continue
        value_text = str(value).strip()
        if value_text and value_text.lower() != "nan":
            return value_text
    return ""


def enrich_company_profile(row, pdl_company_data=None):
    """Fill company enrichment fields without fabricating facts that cannot be verified."""
    row = dict(row or {})
    company_name = _company_field(row, "Company", "Company Name")
    website = _company_field(row, "Website")
    about = _company_field(row, "About Company", "About / Company Description", "About", "Company Description")
    industry = _company_field(row, "Company Industry", "Industry")
    employees = _company_field(row, "No. of Employees", "No of Employees", "Employee Count", "Employees")

    if pdl_company_data and isinstance(pdl_company_data, dict):
        if not about:
            about = _company_field(pdl_company_data, "description", "short_description", "company_description", "about", "overview")
        if not industry:
            industry = _company_field(pdl_company_data, "industry", "industry_name")
            if not industry and isinstance(pdl_company_data.get("industries"), list):
                industry = str(pdl_company_data["industries"][0])
        if not employees:
            employees = _company_field(pdl_company_data, "employee_count", "employee_count_range", "employees")

    if not website:
        website = _company_field(pdl_company_data, "website", "domain") if pdl_company_data else ""

    if company_name and not about:
        about = f"Company profile for {company_name}"
    elif not company_name and website and not about:
        about = f"Company profile from {website}"

    if not about and not company_name and not website:
        about = ""

    if industry and not about:
        about = f"{company_name} operates in the {industry} sector." if company_name else f"Operates in the {industry} sector."

    row["About Company"] = about
    row["Company Industry"] = industry
    row["No. of Employees"] = employees
    row["Website"] = website or row.get("Website", "")
    row["About / Company Description"] = about
    row["Industry"] = industry
    row["Employee Count"] = employees
    return row


def _co_field(row, *keys):
    if not isinstance(row, dict):
        return ""
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return row.get(key)
    return ""


def _get_contact_name_parts(row):
    first = str(_co_field(row, "First Name", "first_name", "firstName") or "").strip()
    last = str(_co_field(row, "Last Name", "last_name", "lastName") or "").strip()

    if (not first or not last) and isinstance(row, dict):
        full_name = str(_co_field(row, "Name", "name", "full_name", "fullName") or "").strip()
        if full_name and re.search(r"\s", full_name):
            parts = full_name.split()
            if not first:
                first = parts[0]
            if not last and len(parts) > 1:
                last = " ".join(parts[1:])

    if not first and last:
        first = last
        last = ""
    return first, last


def _is_demo_profile(profile):
    if not isinstance(profile, dict):
        return False

    needles = []
    for key in ("full_name", "name", "headline"):
        value = profile.get(key)
        if value:
            needles.append(str(value).lower())

    company = profile.get("company") or {}
    if isinstance(company, dict):
        company_name = company.get("name") or ""
    else:
        company_name = str(company)
    if company_name:
        needles.append(str(company_name).lower())

    combined = " ".join(needles)
    demo_markers = [
        "example person",
        "legros, smitham and kessler",
        "manager, business operations & marketing at obm",
        "example-person",
    ]
    return any(marker in combined for marker in demo_markers)


def apollo_search_person(first_name, last_name, company="", title="", include_contacts=False):
    if not APOLLO_API_KEY:
        return {
            "ok": True,
            "found": False,
            "provider": "apollo",
            "profile": None,
            "message": "Apollo API key is not configured."
        }

    payload = {
        "api_key": APOLLO_API_KEY,
        "first_name": first_name,
        "last_name": last_name,
        "page": 1,
        "per_page": 1,
    }

    if company:
        payload["organization_name"] = company
    if title:
        payload["title"] = title

    try:
        response = req.post(
            f"{APOLLO_BASE_URL}/v1/people/search",
            headers={
                "Content-Type": "application/json",
                "x-api-key": APOLLO_API_KEY,
                "Accept": "application/json",
            },
            json=payload,
            timeout=30,
        )

        try:
            data = response.json()
        except ValueError:
            data = {}

        if response.status_code >= 400:
            return {
                "ok": True,
                "found": False,
                "provider": "apollo",
                "profile": None,
                "message": data.get("message", "Apollo request failed."),
                "status_code": response.status_code,
            }

        people = data.get("people") or data.get("results") or []
        if not people:
            return {
                "ok": True,
                "found": False,
                "provider": "apollo",
                "profile": None,
                "message": "No valid Apollo profile found."
            }

        person = people[0]
        company_name = person.get("organization_name") or ((person.get("organization") or {}).get("name")) or company
        headline = person.get("title") or title or ""
        full_name = person.get("name") or f"{first_name} {last_name}".strip()
        result = {
            "ok": True,
            "found": True,
            "provider": "apollo",
            "profile": {
                "first_name": person.get("first_name") or first_name,
                "last_name": person.get("last_name") or last_name,
                "full_name": full_name,
                "headline": headline,
                "url": person.get("linkedin_url") or "",
                "location": person.get("city") or "",
                "country": person.get("country") or "",
                "company": {
                    "name": company_name,
                    "website": person.get("organization_website") or "",
                    "industry": person.get("organization_industry") or "",
                    "size": person.get("organization_size") or "",
                    "overview": "",
                    "headquarter": person.get("organization_city") or "",
                },
                "email": [person.get("email")] if person.get("email") else [],
                "work_email": [person.get("email")] if person.get("email") else [],
                "phone": [person.get("phone")] if person.get("phone") else [],
                "linkedin_url": person.get("linkedin_url") or "",
                "name": full_name,
            }
        }
        if include_contacts:
            result["profile"]["work_email"] = result["profile"]["email"]
        return result

    except req.RequestException as e:
        return {
            "ok": True,
            "found": False,
            "provider": "apollo",
            "profile": None,
            "message": f"Apollo connection error: {str(e)}"
        }

    except Exception as e:
        return {
            "ok": True,
            "found": False,
            "provider": "apollo",
            "profile": None,
            "message": str(e)
        }


def contactout_enrich_person(row, include_contacts=False):
    """
    Enrich one person using ContactOut.

    By default, email/phone are NOT requested
    to avoid unnecessary contact-credit usage.
    """
    first_name, last_name = _get_contact_name_parts(row)
    company_value = _co_field(row, "Company", "company") or ""
    if isinstance(company_value, (list, tuple, set)):
        company = " ".join(str(v).strip() for v in company_value if str(v).strip())
    else:
        company = str(company_value).strip()
    title = str(_co_field(row, "Title", "title", "job_title") or "").strip()
    city = str(_co_field(row, "City", "city", "location") or "").strip()
    website = str(_co_field(row, "Website", "website") or "").strip()

    if not first_name or not last_name:
        return {
            "ok": True,
            "found": False,
            "profile": None,
            "message": "First Name and Last Name are required."
        }

    payload = {
        "first_name": first_name,
        "last_name": last_name,
    }

    if not CONTACTOUT_API_KEY:
        if APOLLO_API_KEY:
            apollo_result = apollo_search_person(first_name, last_name, company, title, include_contacts=include_contacts)
            if apollo_result.get("found"):
                apollo_result["message"] = "ContactOut is not configured; used Apollo fallback."
                return apollo_result
        return {
            "ok": True,
            "found": False,
            "profile": None,
            "message": "ContactOut key is missing or placeholder. Add a real CONTACTOUT_API_KEY to .env."
        }
    if company:
        payload["company"] = [company]

    if title:
        payload["job_title"] = title
        payload["title"] = title

    if city:
        payload["location"] = city

    if website:
        domain = website.strip()
        domain = re.sub(r"^https?://", "", domain, flags=re.I)
        domain = re.sub(r"^www\.", "", domain, flags=re.I)
        domain = domain.split("/")[0]
        if domain:
            payload["company_domain"] = [domain]

    if include_contacts:
        payload["include"] = [
            "work_email",
            "personal_email",
            "phone"
        ]

    try:
        response = req.post(
            f"{CONTACTOUT_BASE_URL}/v1/people/enrich",
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "token": CONTACTOUT_API_KEY
            },
            json=payload,
            timeout=30
        )

        try:
            data = response.json()
        except ValueError:
            data = {}

        if response.status_code >= 400:
            return {
                "ok": True,
                "found": False,
                "provider": "contactout",
                "profile": None,
                "message": data.get("message", "ContactOut request failed."),
                "status_code": response.status_code,
                "details": data,
            }

        profile = data.get("profile")

        if not profile or _is_demo_profile(profile):
            apollo_result = apollo_search_person(first_name, last_name, company, title, include_contacts=include_contacts)
            if apollo_result.get("found"):
                apollo_result["message"] = "No valid ContactOut match found; used Apollo fallback."
                return apollo_result
            if not APOLLO_API_KEY:
                message = "No valid ContactOut profile found and Apollo key is missing or invalid. Add a real APOLLO_API_KEY to .env."
            else:
                message = "No valid ContactOut or Apollo profile found."
            return {
                "ok": True,
                "found": False,
                "provider": "none",
                "profile": None,
                "message": message
            }

        return {
            "ok": True,
            "found": True,
            "provider": "contactout",
            "profile": profile
        }

    except req.RequestException as e:
        return {
            "ok": True,
            "found": False,
            "profile": None,
            "message": f"ContactOut connection error: {str(e)}"
        }

    except Exception as e:
        return {
            "ok": True,
            "found": False,
            "profile": None,
            "message": str(e)
        }


@app.route("/contactout/search", methods=["POST"])
@login_required
def contactout_search():
    data = request.get_json(silent=True) or {}

    full_name = data.get("name") or data.get("full_name") or data.get("fullName") or ""
    if full_name and not (data.get("first_name") or data.get("firstName") or data.get("last_name") or data.get("lastName")):
        parts = full_name.split(None, 1)
        first_name = parts[0] if parts else ""
        last_name = parts[1] if len(parts) > 1 else ""
    else:
        first_name = data.get("first_name") or data.get("firstName") or ""
        last_name = data.get("last_name") or data.get("lastName") or ""

    row = {
        "First Name": first_name,
        "Last Name": last_name,
        "Name": full_name,
        "Company": data.get("company") or "",
        "Title": data.get("title") or data.get("job_title") or data.get("role") or "",
        "City": data.get("city") or data.get("location") or "",
        "Website": data.get("website") or "",
    }

    include_contacts = bool(data.get("include_contacts", False))
    result = contactout_enrich_person(row, include_contacts=include_contacts)
    result.setdefault("found", False)
    return jsonify(result)


@app.route("/contactout/enrich/<int:idx>", methods=["POST"])
@login_required
def contactout_enrich(idx):

    store = get_store()

    if not store:
        return jsonify({"error": "No file loaded"}), 400

    df = store["df"]

    if idx < 0 or idx >= len(df):
        return jsonify({"error": "Invalid row index"}), 400

    row = df.iloc[idx].to_dict()

    data = request.json or {}

    # Default = FALSE
    include_contacts = bool(
        data.get("include_contacts", False)
    )

    result = contactout_enrich_person(
        row,
        include_contacts=include_contacts
    )

    return jsonify(result)

@app.route("/login", methods=["POST"])
def login():
    data = request.json or {}

    username = data.get("username", "").strip().lower()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({
            "error": "Username and password required"
        }), 400

    # Query user from database
    user = User.query.filter_by(username=username).first()

    if user and user.is_active and user.check_password(password):
        session.clear()
        session["username"] = username
        session.permanent = True

        # Update last login
        user.last_login = datetime.now(timezone.utc)
        db.session.commit()

        # Log the login activity
        log_activity(
            user=user,
            action="LOGIN",
            description=f"User {username} logged in"
        )

        return jsonify({
            "ok": True,
            "username": username,
            "role": user.role
        })

    # Log failed login attempt
    log_activity(
        user=username,
        action="FAILED_LOGIN",
        description=f"Failed login attempt for user {username}"
    ) if username else None

    return jsonify({
        "error": "Invalid username or password"
    }), 401

@app.route("/logout", methods=["POST"])
def logout():
    username = session.get("username")
    if username:
        log_activity(
            user=username,
            action="LOGOUT",
            description=f"User {username} logged out"
        )
    session.clear()
    return jsonify({"ok": True})


@app.route("/me")
def me():
    username = session.get("username")
    if not username:
        return jsonify({"logged_in": False})
    
    user = User.query.filter_by(username=username).first()
    if not user or not user.is_active:
        return jsonify({"logged_in": False})
    
    return jsonify({
        "logged_in": True,
        "username": username,
        "role": user.role
    })


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/files")
@login_required
def list_files():
    files = []
    user = get_username()
    prefix = f"{user}_"

    for fname in os.listdir(DRAFT_DIR):
        if not fname.endswith("_draft.xlsx"):
            continue

        if not fname.startswith(prefix):
            continue

        draft_path = os.path.join(DRAFT_DIR, fname)
        base = fname[len(prefix):].replace("_draft.xlsx", "")

        try:
            modified = os.path.getmtime(draft_path)
            modified_str = datetime.fromtimestamp(
                modified
            ).strftime("%b %d, %Y %I:%M %p")

            summary = read_draft_summary(draft_path)
            if summary is None:
                try:
                    df = pd.read_excel(draft_path, dtype=str).fillna("")
                    validated = set(df.index[df["_validated"] == "1"].tolist()) if "_validated" in df.columns else set()
                    summary = build_draft_summary(df, validated)
                    write_draft_summary(draft_path, df, validated)
                except Exception:
                    summary = {"total": 0, "done": 0, "status": "In Progress", "modified_ts": modified}

            total = int(summary.get("total", 0))
            done = int(summary.get("done", 0))
            status = summary.get("status", "In Progress")
            ai_validated = summary.get("ai_validated", False)

            files.append({
                "key": base,
                "filename": base + ".xlsx",
                "total": total,
                "done": done,
                "status": status,
                "modified": modified_str,
                "modified_ts": modified,
                "ai_validated": ai_validated
            })

        except Exception as e:
            print(f"Error reading file info for {fname}: {e}")
            continue

    files.sort(
        key=lambda x: x["modified_ts"],
        reverse=True
    )

    for f in files:
        del f["modified_ts"]

    total_files = len(files)
    completed_files = sum(
        1 for f in files
        if f["status"] == "Completed"
    )
    in_progress_files = total_files - completed_files

    return jsonify({
        "files": files,
        "total_files": total_files,
        "in_progress_files": in_progress_files,
        "completed_files": completed_files
    })


@app.route("/open/<key>")
@login_required
def open_file(key):
    user = get_username()
    if not is_safe_key(key):
        return jsonify({"error": "Invalid file key"}), 400

    draft_path = get_user_draft_path(key)
    if not draft_path or not os.path.exists(draft_path):
        return jsonify({"error": "Draft not found"}), 404

    selected_sheet = request.args.get("sheet")
    selected_validator = request.args.get("validator", "").strip()
    safe_key = os.path.basename(str(key))
    pkl = os.path.join(DRAFT_DIR, f"{user}_{safe_key}.pkl")
    store_key = f"{user}::{key}"

    if os.path.exists(pkl):
        try:
            with open(pkl, "rb") as f:
                store = pickle.load(f)
            stores[store_key] = store
            session["file_key"] = key
            sheet_map = store.get("sheet_map") or {}
            
            # If no sheet selected, check worksheet count
            if not selected_sheet:
                if len(sheet_map) == 1:
                    # Auto-select the only worksheet
                    selected_sheet = next(iter(sheet_map))
                elif len(sheet_map) > 1:
                    # Show selection modal for multiple worksheets
                    return jsonify(build_worksheet_selection_payload(key, sheet_map))
            
            # Handle the selected sheet
            if selected_sheet:
                if selected_sheet not in sheet_map:
                    return jsonify({"error": "Worksheet not found"}), 404
                store["df"] = sheet_map[selected_sheet].copy()
                store["working_sheet_name"] = selected_sheet
                store["masterfile_df"] = None
                store["selected_validator"] = selected_validator or None
                session["file_key"] = key
                
                # Build filtered queue based on validator
                validated, filtered_queue = build_filtered_queue(store["df"], selected_validator)
                store["validated"] = validated
                store["filtered_queue"] = filtered_queue
                store["selected_validator"] = selected_validator or None
                save_store(store_key, store)
                
                resume_index = next((idx for idx in filtered_queue if idx not in validated), filtered_queue[0] if filtered_queue else 0)
                return jsonify({"total": len(store["df"]), "filtered_total": len(filtered_queue), "resume_index": resume_index, "resumed": True, "sheet_name": selected_sheet, "validator": selected_validator})
        except Exception:
            pass

    try:
        workbook_sheets = get_workbook_sheet_map(open(draft_path, "rb"))
        if not workbook_sheets:
            return jsonify({"error": "Could not read Excel file"}), 500

        # If no sheet selected, check worksheet count
        if not selected_sheet:
            if len(workbook_sheets) == 1:
                # Auto-select the only worksheet
                selected_sheet = next(iter(workbook_sheets))
            elif len(workbook_sheets) > 1:
                # Show selection modal for multiple worksheets
                return jsonify(build_worksheet_selection_payload(key, workbook_sheets))

        if selected_sheet:
            if selected_sheet not in workbook_sheets:
                return jsonify({"error": "Worksheet not found"}), 404
            df = workbook_sheets[selected_sheet].copy()
        else:
            return jsonify({"error": "Could not determine worksheet"}), 400

        validated, filtered_queue = build_filtered_queue(df, selected_validator)

        working_sheet_name = selected_sheet or detect_working_sheet_name(workbook_sheets) or next(iter(workbook_sheets))
        store = {
            "df": df,
            "original_df": df.copy(),
            "filename": key + ".xlsx",
            "validated": validated,
            "masterfile_df": None,
            "working_sheet_name": working_sheet_name,
            "sheet_map": workbook_sheets,
            "selected_validator": selected_validator or None,
            "filtered_queue": filtered_queue,
        }
        stores[store_key] = store
        session["file_key"] = key
        save_store(store_key, store)

        resume_index = next((idx for idx in filtered_queue if idx not in validated), filtered_queue[0] if filtered_queue else 0)
        return jsonify({"total": len(df), "filtered_total": len(filtered_queue), "resume_index": resume_index, "resumed": True, "sheet_name": selected_sheet, "validator": selected_validator})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    filename = file.filename
    # For new uploads, use get_draft_path (which auto-increments if needed)
    draft_path = get_draft_path(filename)
    # Extract the key from the path: user_KEY_draft.xlsx -> KEY
    user = get_username()
    draft_base = os.path.basename(draft_path)
    # Format: {user}_{key}_draft.xlsx
    prefix = f"{user}_"
    key = draft_base[len(prefix):].replace("_draft.xlsx", "")

    sheet_map = {}
    df = None
    working_sheet_name = None

    try:
        all_sheets = pd.read_excel(file, sheet_name=None, dtype=str)
        for sheet_name, sheet_df in all_sheets.items():
            sheet_map[sheet_name] = normalize_workbook_sheet(sheet_df.fillna(""))
        working_sheet_name = detect_working_sheet_name(sheet_map) or next(iter(sheet_map), None)
        if working_sheet_name is not None:
            df = sheet_map[working_sheet_name].copy()
    except Exception:
        file.seek(0)
        try:
            df = normalize_workbook_sheet(pd.read_excel(file, dtype=str).fillna(""))
        except Exception:
            return jsonify({"error": "Could not read Excel file"}), 400
        working_sheet_name = "Sheet1"
        sheet_map = {working_sheet_name: df.copy()}

    if df is None:
        return jsonify({"error": "Could not read Excel file"}), 400

    if "be" in df.columns:
        df.rename(columns={"be": "Lead Ranking"}, inplace=True)
    df = normalize_company_fields(df)
    if "Lead Ranking" not in df.columns:
        df["Lead Ranking"] = ""
    if "Validated Date" not in df.columns:
        df["Validated Date"] = ""

    # Detect already-validated rows based on Lead Ranking values
    validated = set()
    for idx, row in df.iterrows():
        if is_row_validated(row):
            validated.add(idx)
    
    resume_index = 0
    if validated:
        # Start from the first unvalidated row
        for i in range(len(df)):
            if i not in validated:
                resume_index = i
                break
        else:
            # All rows validated, start from end
            resume_index = len(df) - 1

    # Store metadata about the workbook structure
    store = {
        "df": df,
        "original_df": df.copy(),
        "filename": filename,
        "validated": validated,
        "masterfile_df": None,
        "working_sheet_name": working_sheet_name or "Sheet1",
        "sheet_map": sheet_map or {working_sheet_name or "Sheet1": df.copy()},
    }
    session["file_key"] = key
    store_key = get_session_key()

    safe_write_excel(df, draft_path, working_sheet_name=store["working_sheet_name"], workbook_sheets=store["sheet_map"])
    write_draft_summary(draft_path, df, validated)
    save_store(store_key, store)

    try:
        uploaded = UploadedFile.query.filter_by(username=user, key=key).first()
        if uploaded is None:
            uploaded = UploadedFile(
                username=user,
                key=key,
                filename=filename,
                file_path=draft_path,
                size_bytes=os.path.getsize(draft_path) if os.path.exists(draft_path) else 0,
            )
            db.session.add(uploaded)
        else:
            uploaded.filename = filename
            uploaded.file_path = draft_path
            uploaded.size_bytes = os.path.getsize(draft_path) if os.path.exists(draft_path) else 0
            uploaded.updated_at = datetime.utcnow()
        db.session.commit()
    except Exception as exc:
        print(f"Failed to persist uploaded file metadata: {exc}")
        db.session.rollback()

    return jsonify({"total": len(df), "columns": list(df.columns), "resume_index": resume_index, "resumed": os.path.exists(draft_path), "user": user})


@app.route("/row/<int:idx>")
@login_required
def get_row(idx):
    store = get_store()
    if not store:
        return jsonify({"error": "No file loaded"}), 400
    
    # Check if row is in filtered queue (if validator filtering is active)
    filtered_queue = store.get("filtered_queue")
    if filtered_queue is not None and idx not in filtered_queue:
        return jsonify({"error": "This row is not in your validation queue"}), 403
    
    df = store["df"]
    if idx >= len(df):
        return jsonify({"done": True})
    row = df.iloc[idx].to_dict()
    validations, suggested_rank, rank_reason = run_validations(row)
    
    filtered_queue = store.get("filtered_queue", list(range(len(df))))
    total_filtered = len(filtered_queue)
    
    return jsonify({"row": row, "index": idx, "total": total_filtered, "validations": validations,
                    "suggested_rank": suggested_rank, "rank_reason": rank_reason,
                    "is_validated": idx in store["validated"]})


@app.route("/progress")
@login_required
def get_progress():
    store = get_store()
    if not store:
        return jsonify({"error": "No file loaded"}), 400
    df = store["df"]
    validated = store["validated"]
    filtered_queue = store.get("filtered_queue", list(range(len(df))))
    selected_validator = store.get("selected_validator")
    
    rows, rows_full = [], []
    
    # Only iterate through filtered queue indices
    for queue_idx, data_idx in enumerate(filtered_queue):
        if data_idx >= len(df):
            continue
        row = df.iloc[data_idx]
        done = data_idx in validated
        rows.append({"index": data_idx, "company": row.get("Company", ""), "rank": row.get("Lead Ranking", ""),
                     "validated_by": row.get("Validated By", ""), "done": done})
        full = row.to_dict()
        full["_index"] = data_idx
        full["_done"] = done
        rows_full.append(full)
    
    done_count = sum(1 for idx in filtered_queue if idx in validated)
    total_filtered = len(filtered_queue)
    
    return jsonify({"rows": rows, "rows_full": rows_full, "total": total_filtered,
                    "done_count": done_count, "validator": selected_validator})


@app.route("/save/<int:idx>", methods=["POST"])
@login_required
def save_row(idx):
    store = get_store()
    if not store:
        return jsonify({"error": "No file loaded"}), 400
    
    # Check if row is in filtered queue (if validator filtering is active)
    filtered_queue = store.get("filtered_queue")
    if filtered_queue is not None and idx not in filtered_queue:
        return jsonify({"error": "This row is not in your validation queue"}), 403
    
    df = store["df"]
    original_df = store.get("original_df")
    data = request.json
    selected_validator = (store.get("selected_validator") or "").strip()

    changed, added, removed = [], [], []
    TRACK_COLS = ["Website", "No. of Employees", "Company Industry", "First Name", "Last Name",
                  "Title", "Email", "Phone", "Alt. Contact Info", "Alternate Phone", "Street", "City", "State", "Zip Code"]
    
    # Track changes for audit log
    changes_dict = {}
    
    if original_df is not None and idx < len(original_df):
        orig_row = original_df.iloc[idx]
        for col in TRACK_COLS:
            new_val = str(data.get(col, "")).strip()
            old_val = str(orig_row.get(col, "")).strip() if col in orig_row else ""
            if old_val and not new_val:
                removed.append(col)
                changes_dict[col] = {"before": old_val, "after": ""}
            elif not old_val and new_val:
                added.append(col)
                changes_dict[col] = {"before": "", "after": new_val}
            elif old_val and new_val and old_val != new_val:
                changed.append(col)
                changes_dict[col] = {"before": old_val, "after": new_val}

    parts = []
    if changed: parts.append("Changes: " + ", ".join(changed))
    if removed: parts.append("Removed: " + ", ".join(removed))
    if added:   parts.append("Added: " + ", ".join(added))
    changes_text = " | ".join(parts)

    normalized_payload = {}
    for key, value in data.items():
        if key == "Validated Date":
            normalized_payload[key] = normalize_validated_date(value)
        elif key == "Lead Ranking":
            normalized_payload[key] = normalize_lead_ranking_value(value)
        else:
            normalized_payload[key] = value

    for key in ["Company", "Validated By", "Validated Date"]:
        if key not in df.columns:
            df[key] = ""

    if selected_validator:
        normalized_payload["Validated By"] = selected_validator

    for key, value in normalized_payload.items():
        if key in df.columns:
            df.at[idx, key] = value

    if "Notes" in df.columns and changes_text:
        lines = [l for l in str(df.at[idx, "Notes"]).strip().split("\n")
                 if not l.startswith(("Changes:", "Added:", "Removed:"))]
        lines.append(changes_text)
        df.at[idx, "Notes"] = "\n".join(lines).strip()

    store["validated"].add(idx)
    store["df"] = df
    if store.get("sheet_map"):
        selected_name = store.get("working_sheet_name")
        if selected_name and selected_name in store["sheet_map"]:
            store["sheet_map"][selected_name] = df.copy()

    key = get_session_key()
    file_key = session.get("file_key")
    if not file_key:
        return jsonify({"error": "No file loaded"}), 400

    flush_store_to_disk(store, file_key=file_key)
    save_store(key, store)

    navigation_queue = filtered_queue if filtered_queue is not None else list(range(len(df)))
    next_index = None
    current_queue_pos = navigation_queue.index(idx)
    if current_queue_pos + 1 < len(navigation_queue):
        next_index = navigation_queue[current_queue_pos + 1]

    # Log the save activity
    company_name = data.get("Company", "Unknown")
    log_activity(
        user=get_current_user(),
        action="SAVE_LEAD",
        description=f"Saved lead: {company_name}",
        lead_id=file_key,
        changes=changes_dict if changes_dict else None
    )

    return jsonify({"ok": True, "done_count": sum(1 for i in store["validated"] if i < len(df)),
                    "changes_text": changes_text, "next_index": next_index})


@app.route("/delete/<int:idx>", methods=["POST"])
@login_required
def delete_row(idx):
    store = get_store()
    if not store:
        return jsonify({"error": "No file loaded"}), 400
    
    # Get company name before deletion for audit log
    company_name = store["df"].iloc[idx].get("Company", "Unknown") if idx < len(store["df"]) else "Unknown"
    file_key = session.get("file_key", "Unknown")
    
    df = store["df"].drop(index=idx).reset_index(drop=True)
    store["df"] = df
    if store.get("original_df") is not None:
        store["original_df"] = store["original_df"].drop(index=idx).reset_index(drop=True)
    store["validated"] = {i if i < idx else i - 1 for i in store["validated"] if i != idx}

    key = get_session_key()
    if not file_key:
        return jsonify({"error": "No file loaded"}), 400

    flush_store_to_disk(store, file_key=file_key)
    save_store(key, store)

    return jsonify({"ok": True, "total": len(df)})


@app.route("/download")
@login_required
def download():
    store = get_store()
    if not store:
        return jsonify({"error": "No file loaded"}), 400
    df = store["df"]
    masterfile_df = store.get("masterfile_df")
    working_sheet_name = store.get("working_sheet_name") or "Sheet1"
    workbook_sheets = store.get("sheet_map")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if workbook_sheets:
            for sheet_name, sheet_df in workbook_sheets.items():
                if sheet_df is not None:
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            if masterfile_df is not None:
                masterfile_df.to_excel(writer, sheet_name="Masterfile", index=False)
            df.to_excel(writer, sheet_name=working_sheet_name, index=False)
    output.seek(0)
    filename = store["filename"].replace(".xlsx", "_validated.xlsx")
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/download/<key>")
@login_required
def download_file_by_key(key):
    """Download a specific file from the Files system."""
    user = get_username()
    if not is_safe_key(key):
        return jsonify({"error": "Invalid file key"}), 400
    
    draft_path = get_user_draft_path(key)
    if not draft_path or not os.path.exists(draft_path):
        return jsonify({"error": "File not found"}), 404
    
    # Verify the file belongs to the current user
    expected_prefix = f"{user}_"
    filename = os.path.basename(draft_path)
    if not filename.startswith(expected_prefix):
        return jsonify({"error": "Unauthorized"}), 403
    
    # Generate a friendly download name
    display_name = key.replace("_draft", "").replace("_", " ") + ".xlsx"
    
    return send_file(
        draft_path,
        download_name=display_name,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/delete/<key>", methods=["POST"])
@login_required
def delete_uploaded_file(key):
    """Delete a user's draft file and its metadata after confirmation."""
    user = get_username()
    if not user:
        return jsonify({"error": "Not logged in"}), 401

    if not is_safe_key(key):
        return jsonify({"error": "Invalid file key"}), 400

    safe_key = os.path.basename(str(key))
    draft_path = get_user_draft_path(safe_key)
    if not draft_path:
        return jsonify({"error": "Invalid file key"}), 400

    try:
        db_record = UploadedFile.query.filter_by(username=user, key=safe_key).first()
        if db_record is not None:
            db.session.delete(db_record)

        for candidate in [draft_path, os.path.join(DRAFT_DIR, f"{user}_{safe_key}.pkl")]:
            if candidate and os.path.exists(candidate):
                os.remove(candidate)

        summary_path = get_draft_summary_path(draft_path)
        if os.path.exists(summary_path):
            os.remove(summary_path)

        if db_record is not None and db_record.file_path and os.path.exists(db_record.file_path) and os.path.abspath(db_record.file_path) != os.path.abspath(draft_path):
            os.remove(db_record.file_path)

        db.session.commit()

        log_activity(
            user=user,
            action="DELETE_FILE",
            description=f"Deleted file: {safe_key}",
            lead_id=safe_key
        )

        return jsonify({"ok": True, "message": "File deleted successfully."})
    except Exception as exc:
        db.session.rollback()
        print(f"Error deleting file {safe_key}: {exc}")
        return jsonify({"error": "Unable to delete file. Please try again."}), 500


# ============================================================================
# ADMIN DASHBOARD ROUTES
# ============================================================================

@app.route("/admin")
@admin_required
def admin_dashboard():
    """Admin Dashboard with statistics."""
    try:
        total_users = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()
        inactive_users = total_users - active_users
        
        # Count leads (files) from drafts directory
        total_leads = 0
        validated_leads = 0
        
        for fname in os.listdir(DRAFT_DIR):
            if fname.endswith("_draft.xlsx"):
                total_leads += 1
                try:
                    df = pd.read_excel(os.path.join(DRAFT_DIR, fname), dtype=str).fillna("")
                    if "_validated" in df.columns:
                        validated_leads += sum(1 for v in df["_validated"] if v == "1")
                except Exception:
                    pass
        
        # Recent activity (last 10 logs)
        recent_activity = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()
        recent_activity_data = [log.to_dict() for log in recent_activity]
        
        return render_template("admin/dashboard.html", 
                             total_users=total_users,
                             active_users=active_users,
                             inactive_users=inactive_users,
                             total_leads=total_leads,
                             validated_leads=validated_leads,
                             recent_activity=recent_activity_data)
    except Exception as e:
        print(f"Error rendering admin dashboard: {e}")
        return jsonify({"error": "Error loading dashboard"}), 500


@app.route("/admin/users")
@admin_required
def admin_users():
    """User Management page."""
    try:
        page = request.args.get("page", 1, type=int)
        per_page = 10
        
        users_paginated = User.query.order_by(User.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        users = [user.to_dict() for user in users_paginated.items]
        
        return render_template("admin/users.html",
                             users=users,
                             page=page,
                             total_pages=users_paginated.pages,
                             total_users=users_paginated.total)
    except Exception as e:
        print(f"Error rendering user management: {e}")
        return jsonify({"error": "Error loading users"}), 500


@app.route("/admin/logs")
@admin_required
def admin_logs():
    """Activity Logs page."""
    try:
        page = request.args.get("page", 1, type=int)
        per_page = 20
        user_filter = request.args.get("user", "")
        action_filter = request.args.get("action", "")
        
        query = ActivityLog.query
        
        if user_filter:
            user = User.query.filter_by(username=user_filter).first()
            if user:
                query = query.filter_by(user_id=user.id)
        
        if action_filter:
            query = query.filter_by(action=action_filter)
        
        logs_paginated = query.order_by(ActivityLog.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        logs = [log.to_dict() for log in logs_paginated.items]
        
        # Get unique actions and users for filters
        all_actions = db.session.query(ActivityLog.action).distinct().all()
        actions = [a[0] for a in all_actions]
        
        all_users = User.query.all()
        usernames = [u.username for u in all_users]
        
        return render_template("admin/logs.html",
                             logs=logs,
                             page=page,
                             total_pages=logs_paginated.pages,
                             total_logs=logs_paginated.total,
                             user_filter=user_filter,
                             action_filter=action_filter,
                             available_actions=actions,
                             available_users=usernames)
    except Exception as e:
        print(f"Error rendering activity logs: {e}")
        return jsonify({"error": "Error loading logs"}), 500


# ============================================================================
# ADMIN API ROUTES (for user management operations)
# ============================================================================

@app.route("/api/admin/users", methods=["POST"])
@admin_required
def api_create_user():
    """Create a new user."""
    try:
        data = request.json or {}
        username = data.get("username", "").strip().lower()
        password = data.get("password", "").strip()
        role = data.get("role", "User")
        
        if not username or not password:
            return jsonify({"error": "Username and password required"}), 400
        
        if role not in ["Admin", "User", "Viewer"]:
            return jsonify({"error": "Invalid role"}), 400
        
        # Check if user already exists
        existing = User.query.filter_by(username=username).first()
        if existing:
            return jsonify({"error": "User already exists"}), 409
        
        # Create new user
        user = User(username=username, role=role, is_active=True)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        # Log the activity
        current_user = get_current_user()
        log_activity(
            user=current_user,
            action="CREATE_USER",
            description=f"Created user {username} with role {role}"
        )
        
        return jsonify({"ok": True, "user": user.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        print(f"Error creating user: {e}")
        return jsonify({"error": "Error creating user"}), 500


@app.route("/api/admin/users/<int:user_id>", methods=["PUT"])
@admin_required
def api_update_user(user_id):
    """Update an existing user."""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        data = request.json or {}
        
        if "role" in data:
            role = data.get("role")
            if role not in ["Admin", "User", "Viewer"]:
                return jsonify({"error": "Invalid role"}), 400
            user.role = role
            
            # Log role change
            current_user = get_current_user()
            log_activity(
                user=current_user,
                action="CHANGE_ROLE",
                description=f"Changed {user.username}'s role to {role}"
            )
        
        if "password" in data:
            password = data.get("password", "").strip()
            if password:
                user.set_password(password)
                
                # Log password reset
                current_user = get_current_user()
                log_activity(
                    user=current_user,
                    action="RESET_PASSWORD",
                    description=f"Reset password for user {user.username}"
                )
        
        user.updated_at = datetime.now(timezone.utc)
        db.session.commit()
        
        return jsonify({"ok": True, "user": user.to_dict()})
    except Exception as e:
        db.session.rollback()
        print(f"Error updating user: {e}")
        return jsonify({"error": "Error updating user"}), 500


@app.route("/api/admin/users/<int:user_id>/deactivate", methods=["POST"])
@admin_required
def api_deactivate_user(user_id):
    """Deactivate a user."""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        # Prevent self-deactivation
        current_user = get_current_user()
        if user.id == current_user.id:
            return jsonify({"error": "Cannot deactivate yourself"}), 400
        
        user.is_active = False
        user.updated_at = datetime.now(timezone.utc)
        db.session.commit()
        
        # Log the activity
        log_activity(
            user=current_user,
            action="DEACTIVATE_USER",
            description=f"Deactivated user {user.username}"
        )
        
        return jsonify({"ok": True, "user": user.to_dict()})
    except Exception as e:
        db.session.rollback()
        print(f"Error deactivating user: {e}")
        return jsonify({"error": "Error deactivating user"}), 500


@app.route("/api/admin/users/<int:user_id>/reactivate", methods=["POST"])
@admin_required
def api_reactivate_user(user_id):
    """Reactivate a user."""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        user.is_active = True
        user.updated_at = datetime.now(timezone.utc)
        db.session.commit()
        
        # Log the activity
        current_user = get_current_user()
        log_activity(
            user=current_user,
            action="REACTIVATE_USER",
            description=f"Reactivated user {user.username}"
        )
        
        return jsonify({"ok": True, "user": user.to_dict()})
    except Exception as e:
        db.session.rollback()
        print(f"Error reactivating user: {e}")
        return jsonify({"error": "Error reactivating user"}), 500


# ============================================================================
# AI VALIDATION ROUTES
# ============================================================================

# Store for AI validation job progress: {job_id: {status, progress_data}}
ai_validation_jobs = {}


def pdl_enrich_person(first_name, last_name, company="", title=""):
    """Enrich person data using People Data Labs API."""
    if not PDL_API_KEY:
        return {"found": False, "message": "PDL API key not configured"}
    
    try:
        headers = {
            "X-Api-Key": PDL_API_KEY,
            "Content-Type": "application/json"
        }
        
        payload = {
            "query": {
                "bool": {
                    "must": []
                }
            }
        }
        
        # Build query with first and last name
        if first_name:
            payload["query"]["bool"]["must"].append({
                "match": {
                    "first_name": {
                        "query": first_name,
                        "fuzziness": "AUTO"
                    }
                }
            })
        
        if last_name:
            payload["query"]["bool"]["must"].append({
                "match": {
                    "last_name": {
                        "query": last_name,
                        "fuzziness": "AUTO"
                    }
                }
            })
        
        if company:
            payload["query"]["bool"]["must"].append({
                "match": {
                    "work_company_name": {
                        "query": company,
                        "fuzziness": "AUTO"
                    }
                }
            })
        
        if title:
            payload["query"]["bool"]["must"].append({
                "match": {
                    "job_title": {
                        "query": title,
                        "fuzziness": "AUTO"
                    }
                }
            })
        
        response = req.post(
            f"{PDL_BASE_URL}/person/search",
            headers=headers,
            json=payload,
            timeout=15
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get("data") and len(data["data"]) > 0:
                person = data["data"][0]
                return {
                    "found": True,
                    "person": person,
                    "message": "Found on PDL"
                }
        
        return {
            "found": False,
            "message": f"PDL query returned status {response.status_code}"
        }
    
    except Exception as e:
        return {
            "found": False,
            "message": f"PDL error: {str(e)}"
        }


def pdl_enrich_company(company_name, website=""):
    """Enrich company data using People Data Labs API."""
    if not PDL_API_KEY:
        return {"found": False, "message": "PDL API key not configured"}
    
    try:
        headers = {
            "X-Api-Key": PDL_API_KEY,
            "Content-Type": "application/json"
        }
        
        payload = {
            "query": {
                "bool": {
                    "must": []
                }
            }
        }
        
        if company_name:
            payload["query"]["bool"]["must"].append({
                "match": {
                    "name": {
                        "query": company_name,
                        "fuzziness": "AUTO"
                    }
                }
            })
        
        if website:
            domain = website.strip()
            domain = re.sub(r"^https?://", "", domain, flags=re.I)
            domain = re.sub(r"^www\.", "", domain, flags=re.I)
            domain = domain.split("/")[0]
            if domain:
                payload["query"]["bool"]["must"].append({
                    "term": {
                        "website": domain
                    }
                })
        
        response = req.post(
            f"{PDL_BASE_URL}/company/search",
            headers=headers,
            json=payload,
            timeout=15
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get("data") and len(data["data"]) > 0:
                company = data["data"][0]
                return {
                    "found": True,
                    "company": company,
                    "message": "Found on PDL"
                }
        
        return {
            "found": False,
            "message": f"PDL query returned status {response.status_code}"
        }
    
    except Exception as e:
        return {
            "found": False,
            "message": f"PDL error: {str(e)}"
        }

@app.route("/api/ai-validation/validators")
@login_required
def get_validators():
    """Get list of validators from existing Excel files."""
    try:
        validators = set()
        user = get_username()
        prefix = f"{user}_"
        
        # Scan all user's draft files for "Validated By" column values
        for fname in os.listdir(DRAFT_DIR):
            if not fname.endswith("_draft.xlsx") or not fname.startswith(prefix):
                continue
            
            draft_path = os.path.join(DRAFT_DIR, fname)
            try:
                df = pd.read_excel(draft_path, dtype=str).fillna("")
                if "Validated By" in df.columns:
                    for val in df["Validated By"]:
                        if val and str(val).strip():
                            validators.add(str(val).strip())
            except Exception:
                pass
        
        return jsonify({"validators": sorted(list(validators))})
    except Exception as e:
        print(f"Error loading validators: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/ai-validation/inspect-worksheets", methods=["POST"])
@login_required
def inspect_worksheets():
    """Inspect uploaded workbook and return worksheet information."""
    print("[AI VALIDATION] ========== INSPECT WORKSHEETS CALLED ==========")
    try:
        file = request.files.get("file")
        
        print(f"[AI VALIDATION] File: {file.filename if file else 'NONE'}")
        
        if not file:
            print("[AI VALIDATION] ERROR: No file uploaded")
            return jsonify({"error": "No file uploaded"}), 400
        
        # Read the workbook
        try:
            print(f"[AI VALIDATION] Reading workbook from file...")
            workbook_sheets = get_workbook_sheet_map(file)
            print(f"[AI VALIDATION] Workbook sheets: {list(workbook_sheets.keys()) if workbook_sheets else 'NONE'}")
            if not workbook_sheets:
                print("[AI VALIDATION] ERROR: Could not read Excel file (no sheets)")
                return jsonify({"error": "Could not read Excel file"}), 400
        except Exception as e:
            print(f"[AI VALIDATION] ERROR reading workbook: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Could not read Excel file: {str(e)}"}), 400
        
        # Detect working sheet
        working_sheet_name = detect_working_sheet_name(workbook_sheets) or next(iter(workbook_sheets), None)
        print(f"[AI VALIDATION] Working sheet detected: {working_sheet_name}")
        
        # Analyze each worksheet
        worksheets_info = []
        for sheet_name, df in workbook_sheets.items():
            if df is None or len(df) == 0:
                continue
            
            total_rows = int(len(df))  # Convert to int
            
            # Count already validated rows
            validated_by_column = None
            for col in df.columns:
                if col.lower() in ['validated by', 'validatedby']:
                    validated_by_column = col
                    break
            
            already_validated = 0
            if validated_by_column:
                already_validated = int(df[validated_by_column].notna().sum() - (df[validated_by_column] == '').sum() - (df[validated_by_column].isnull()).sum())
            else:
                already_validated = 0
            
            needs_validation = int(total_rows - already_validated)  # Convert to int
            
            worksheets_info.append({
                'name': sheet_name,
                'total': total_rows,
                'already_validated': already_validated,
                'needs_validation': needs_validation,
                'is_working': sheet_name == working_sheet_name,
                'validated_by_counts': {}
            })
            
            # Count validators
            if validated_by_column:
                validator_counts = df[validated_by_column].value_counts().to_dict()
                worksheets_info[-1]['validated_by_counts'] = {
                    str(k): int(v) for k, v in validator_counts.items() if k and str(k).strip() and k != 'nan'
                }
            
            print(f"[AI VALIDATION] Sheet '{sheet_name}': {total_rows} total, {already_validated} validated, {needs_validation} need validation")
        
        print(f"[AI VALIDATION] Returning info for {len(worksheets_info)} worksheets")
        return jsonify({
            "worksheets": worksheets_info,
            "working_sheet": working_sheet_name
        })
    
    except Exception as e:
        print(f"[AI VALIDATION] EXCEPTION in inspect_worksheets: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/ai-validation/start", methods=["POST"])
@login_required
def start_ai_validation():
    """Start an AI validation job."""
    print("[AI VALIDATION] ========== START ROUTE CALLED ==========")
    try:
        file = request.files.get("file")
        validator = request.form.get("validator", "").strip()
        selected_worksheet = request.form.get("worksheet", "").strip()
        
        print(f"[AI VALIDATION] File: {file.filename if file else 'NONE'}")
        print(f"[AI VALIDATION] Validator: {validator}")
        print(f"[AI VALIDATION] Selected Worksheet: {selected_worksheet}")
        
        if not file:
            print("[AI VALIDATION] ERROR: No file uploaded")
            return jsonify({"error": "No file uploaded"}), 400
        if not validator:
            print("[AI VALIDATION] ERROR: No validator specified")
            return jsonify({"error": "Validator required"}), 400
        
        # Generate job ID
        job_id = f"{get_username()}_{int(time.time())}_{os.urandom(4).hex()}"
        print(f"[AI VALIDATION] Generated job_id: {job_id}")
        
        # Read the workbook
        try:
            print(f"[AI VALIDATION] Reading workbook from file...")
            workbook_sheets = get_workbook_sheet_map(file)
            print(f"[AI VALIDATION] Workbook sheets: {list(workbook_sheets.keys()) if workbook_sheets else 'NONE'}")
            if not workbook_sheets:
                print("[AI VALIDATION] ERROR: Could not read Excel file (no sheets)")
                return jsonify({"error": "Could not read Excel file"}), 400
        except Exception as e:
            print(f"[AI VALIDATION] ERROR reading workbook: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Could not read Excel file: {str(e)}"}), 400
        
        # Use selected worksheet or detect working sheet
        if selected_worksheet and selected_worksheet in workbook_sheets:
            working_sheet_name = selected_worksheet
            print(f"[AI VALIDATION] Using selected worksheet: {working_sheet_name}")
        else:
            working_sheet_name = detect_working_sheet_name(workbook_sheets) or next(iter(workbook_sheets), "Sheet1")
            print(f"[AI VALIDATION] Detected working sheet: {working_sheet_name}")
        
        df = workbook_sheets.get(working_sheet_name)
        
        if df is None or len(df) == 0:
            print(f"[AI VALIDATION] ERROR: No data in working sheet (df={df}, len={len(df) if df is not None else 'None'})")
            return jsonify({"error": "No data in workbook"}), 400
        
        print(f"[AI VALIDATION] Dataframe loaded: {len(df)} rows, {len(df.columns)} columns")
        
        # Initialize job
        ai_validation_jobs[job_id] = {
            "status": "processing",
            "user": get_username(),
            "validator": validator,
            "original_filename": file.filename,  # Store original filename for output naming
            "df": df.copy(),
            "workbook_sheets": workbook_sheets,
            "working_sheet_name": working_sheet_name,
            "total": len(df),
            "eligible_rows": [],  # Will be calculated in process_ai_validation_async
            "eligible_count": 0,  # Count of rows to process
            "processed": 0,
            "verified": 0,
            "partial": 0,
            "conflict": 0,
            "not_found": 0,
            "errors": 0,
            "current_status": "Initializing...",
            "created_at": datetime.now(),
            "download_path": None,
            "file_key": None  # Will be set after save
        }
        
        print(f"[AI VALIDATION] Job initialized: {job_id}")
        print(f"[AI VALIDATION] Starting background processing...")
        
        # Start background processing asynchronously so the browser can receive the
        # job_id and begin polling immediately instead of waiting for the full workbook
        # validation to finish in the same request.
        thread = threading.Thread(target=process_ai_validation_async, args=(job_id,), daemon=True)
        thread.start()
        print(f"[AI VALIDATION] Background thread started for job: {job_id}")
        
        print(f"[AI VALIDATION] Returning success response with job_id: {job_id}")
        return jsonify({"ok": True, "job_id": job_id})
    
    except Exception as e:
        print(f"[AI VALIDATION] EXCEPTION in start_ai_validation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def process_ai_validation_async(job_id):
    """Process AI validation in background."""
    print(f"[AI VALIDATION] process_ai_validation_async STARTED for job: {job_id}")
    try:
        job = ai_validation_jobs.get(job_id)
        if not job:
            print(f"[AI VALIDATION] ERROR: Job {job_id} not found in ai_validation_jobs")
            return
        
        print(f"[AI VALIDATION] Job found: status={job['status']}, total={job['total']}")
        
        df = job["df"].copy()
        validator = job["validator"]
        print(f"[AI VALIDATION] DataFrame copied, {len(df)} rows, validator={validator}")
        
        # Ensure required columns exist
        ensure_company_enrichment_columns(df)
        if "Validated By" not in df.columns:
            df["Validated By"] = ""
        if "Validated Date" not in df.columns:
            df["Validated Date"] = ""
        if "Validation Status" not in df.columns:
            df["Validation Status"] = ""
        if "Lead Ranking" not in df.columns:
            df["Lead Ranking"] = ""
        if "Notes" not in df.columns:
            df["Notes"] = ""
        
        print(f"[AI VALIDATION] Required columns initialized")
        
        # FILTER: Identify eligible rows for processing
        # Eligible rows are those where:
        # 1. Validated By equals the selected validator (must be assigned to them)
        # 2. AND Lead Ranking is blank (not Bad/Good/Better/Best)
        eligible_rows = []
        completed_count = 0
        assigned_to_others = 0
        not_assigned = 0
        
        for idx in range(len(df)):
            row = df.iloc[idx]
            validated_by = str(row.get("Validated By", "")).strip()
            lead_ranking = str(row.get("Lead Ranking", "")).strip()
            
            # Check if row is already completed
            if lead_ranking and is_completed_lead_ranking(lead_ranking):
                completed_count += 1
                continue
            
            # Check if row is not assigned to this validator
            if validated_by != validator:
                # It's either assigned to someone else or not assigned at all
                if validated_by:
                    assigned_to_others += 1
                else:
                    not_assigned += 1
                continue
            
            # This row is eligible for processing (assigned to this validator with blank Lead Ranking)
            eligible_rows.append(idx)
        
        job["eligible_rows"] = eligible_rows
        job["eligible_count"] = len(eligible_rows)
        
        print(f"[AI VALIDATION] Filtering complete:")
        print(f"  Total rows: {len(df)}")
        print(f"  Eligible for {validator}: {len(eligible_rows)}")
        print(f"  Already completed: {completed_count}")
        print(f"  Assigned to other validators: {assigned_to_others}")
        print(f"  Not assigned yet: {not_assigned}")
        
        # Process only eligible rows
        print(f"[AI VALIDATION] Starting row processing: {len(eligible_rows)} eligible rows")
        for row_num, idx in enumerate(eligible_rows):
            row = df.iloc[idx]
            
            try:
                company_name = str(row.get("Company", "")).strip()
                first_name = str(row.get("First Name", "")).strip()
                last_name = str(row.get("Last Name", "")).strip()
                title = str(row.get("Title", "")).strip()
                website = str(row.get("Website", "")).strip()
                
                job["current_status"] = f"Validating: {company_name or f'{first_name} {last_name}'} ({job['processed'] + 1}/{len(eligible_rows)})"
                
                # Step 1: Run existing validation functions on the current row
                validations, suggested_rank, rank_reason = run_validations(row)
                
                # Step 2: Automatically enrich the company and contact details using company data
                pdl_company_data = None
                pdl_person_data = None
                
                if company_name:
                    pdl_company_result = pdl_enrich_company(company_name, website)
                    if pdl_company_result.get("found"):
                        pdl_company_data = pdl_company_result.get("company", {})
                        enriched_row = enrich_company_profile(row.to_dict(), pdl_company_data)
                        for key, value in enriched_row.items():
                            if value and (key not in df.columns or not str(df.at[idx, key]).strip()):
                                if key in df.columns:
                                    df.at[idx, key] = value
                        if not str(df.at[idx, "Website"]).strip() and pdl_company_data.get("website"):
                            df.at[idx, "Website"] = pdl_company_data["website"]
                        if not str(df.at[idx, "No. of Employees"]).strip() and pdl_company_data.get("employee_count"):
                            df.at[idx, "No. of Employees"] = str(pdl_company_data["employee_count"])
                        if not str(df.at[idx, "Company Industry"]).strip() and pdl_company_data.get("industry"):
                            df.at[idx, "Company Industry"] = pdl_company_data["industry"]
                        if not str(df.at[idx, "About Company"]).strip() and pdl_company_data.get("description"):
                            df.at[idx, "About Company"] = pdl_company_data["description"]
                
                if first_name and last_name:
                    pdl_person_result = pdl_enrich_person(first_name, last_name, company_name, title)
                    if pdl_person_result.get("found"):
                        pdl_person_data = pdl_person_result.get("person", {})
                        
                        # Enrich person fields from PDL
                        if not str(row.get("Title", "")).strip() and pdl_person_data.get("job_title"):
                            df.at[idx, "Title"] = pdl_person_data["job_title"]
                        
                        if not str(row.get("Email", "")).strip() and pdl_person_data.get("emails"):
                            emails = pdl_person_data.get("emails", [])
                            if emails and len(emails) > 0:
                                df.at[idx, "Email"] = emails[0]
                        
                        if not str(row.get("Phone", "")).strip() and pdl_person_data.get("phone_numbers"):
                            phones = pdl_person_data.get("phone_numbers", [])
                            if phones and len(phones) > 0:
                                df.at[idx, "Phone"] = phones[0]
                
                # Step 3-6: Re-run validations and recalculate ranking after enrichment
                updated_row = df.iloc[idx].to_dict()
                validations, suggested_rank, rank_reason = run_validations(updated_row)
                if suggested_rank.lower() == "manual review":
                    df.at[idx, "Manual Review"] = "Needs manual review"
                else:
                    df.at[idx, "Manual Review"] = ""
                
                # Determine validation status based on validations
                all_valid = all(v.get("valid", False) for v in validations.values())
                some_valid = any(v.get("valid", False) for v in validations.values())
                
                if all_valid:
                    validation_status = "Verified"
                    job["verified"] += 1
                elif some_valid:
                    validation_status = "Partial"
                    job["partial"] += 1
                else:
                    validation_status = "Not Found"
                    job["not_found"] += 1
                
                # Build notes explaining what was verified/found
                notes_parts = []
                
                # Add what was verified
                verified_fields = [field for field, result in validations.items() if result.get("valid", False)]
                if verified_fields:
                    notes_parts.append(f"Verified: {', '.join(verified_fields)}")
                
                # Add what failed
                failed_fields = [field for field, result in validations.items() if not result.get("valid", False)]
                if failed_fields:
                    notes_parts.append(f"Not verified: {', '.join(failed_fields)}")
                
                # Add enrichment notes
                if pdl_company_data or pdl_person_data:
                    notes_parts.append("Enriched from PDL")
                
                # Update row with validation results
                df.at[idx, "Validation Status"] = validation_status
                # Validated By is already set to the validator (it's a filter requirement)
                # Only update other fields: Validated Date, Lead Ranking, Notes
                df.at[idx, "Validated Date"] = datetime.now().strftime("%b %d, %Y")
                df.at[idx, "Lead Ranking"] = suggested_rank
                if notes_parts:
                    df.at[idx, "Notes"] = " | ".join(notes_parts)
                if suggested_rank.lower() == "manual review":
                    df.at[idx, "Manual Review"] = "Needs manual review"
                else:
                    df.at[idx, "Manual Review"] = ""
                
                job["processed"] += 1
                
            except Exception as e:
                job["errors"] += 1
                job["processed"] += 1
                df.at[idx, "Validation Status"] = "Error"
                existing_notes = str(df.at[idx, "Notes"]).strip()
                error_msg = f"Validation error: {str(e)}"
                if existing_notes:
                    df.at[idx, "Notes"] = f"{existing_notes} | {error_msg}"
                else:
                    df.at[idx, "Notes"] = error_msg
                print(f"Error processing row {idx}: {e}")
                continue
            
            # Small delay to be respectful to APIs
            time.sleep(0.1)
        
        # Save the validated workbook and register it in the Files system
        try:
            # Generate a unique file_key for the output based on original filename
            original_filename = job.get("original_filename", "ai_validated")
            base_name = os.path.splitext(original_filename)[0]
            # Clean the base name: remove .xlsx if present, limit length
            base_name = base_name.replace(" ", "_").replace(".", "_")[:50]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ai_output_key = f"ai_validated_{base_name}_{timestamp}"

            job_user = str(job.get("user") or get_username() or "").strip()
            draft_path = build_user_draft_path(job_user, ai_output_key)
            if not draft_path:
                raise RuntimeError("Cannot determine destination path for validated workbook")

            # Update the working sheet in the workbook
            workbook_sheets = job["workbook_sheets"]
            if job["working_sheet_name"] in workbook_sheets:
                workbook_sheets[job["working_sheet_name"]] = df.copy()

            safe_write_excel(
                df,
                draft_path,
                working_sheet_name=job["working_sheet_name"],
                workbook_sheets=workbook_sheets
            )
            if not os.path.exists(draft_path):
                raise FileNotFoundError(f"Validated workbook was not created: {draft_path}")

            apply_lead_ranking_conditional_formatting(draft_path)

            # Persist metadata for the Files page and download route.
            try:
                uploaded = UploadedFile.query.filter_by(username=job_user, key=ai_output_key).first()
                if uploaded is None:
                    uploaded = UploadedFile(
                        username=job_user,
                        key=ai_output_key,
                        filename=os.path.basename(draft_path),
                        file_path=draft_path,
                        size_bytes=os.path.getsize(draft_path),
                    )
                    db.session.add(uploaded)
                else:
                    uploaded.filename = os.path.basename(draft_path)
                    uploaded.file_path = draft_path
                    uploaded.size_bytes = os.path.getsize(draft_path)
                    uploaded.updated_at = datetime.utcnow()
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                print(f"[AI VALIDATION] Warning: could not persist UploadedFile metadata: {exc}")

            # Create draft summary with AI validation metadata
            # Count rows with Lead Ranking filled (validated)
            validated_count = 0
            for idx, row in df.iterrows():
                lead_ranking = str(row.get("Lead Ranking", "")).strip()
                if lead_ranking and lead_ranking.lower() in ["bad", "good", "better", "best"]:
                    validated_count += 1

            # Create extended summary with AI validation info
            ai_summary = {
                "total": len(df),
                "done": validated_count,
                "status": "Completed",
                "modified_ts": time.time(),
                "ai_validated": True,
                "validator": job["validator"],
                "eligible_count": job.get("eligible_count", len(df)),
                "processed": job["processed"],
                "verified": job.get("verified", 0),
                "partial": job.get("partial", 0),
                "not_found": job.get("not_found", 0),
                "errors": job.get("errors", 0),
            }
            summary_path = get_draft_summary_path(draft_path)
            try:
                with open(summary_path, "w", encoding="utf-8") as f:
                    json.dump(ai_summary, f)
            except Exception as e:
                print(f"[AI VALIDATION] Warning: Could not write draft summary: {e}")

            # Store the file_key and path in the job for retrieval.
            job["download_path"] = draft_path
            job["file_key"] = ai_output_key
            job["df"] = df

            print(f"[AI VALIDATION] Workbook saved to: {draft_path}")
            print(f"[AI VALIDATION] File key: {ai_output_key}")
            print(f"[AI VALIDATION] File registered in Files system")

        except Exception as e:
            print(f"[AI VALIDATION] ERROR saving validated workbook: {e}")
            import traceback
            traceback.print_exc()
            job["errors"] += 1
            job["status"] = "error"
            job["current_status"] = f"Error: {str(e)}"
            print(f"[AI VALIDATION] JOB FAILED: {job_id}")
            print(f"[AI VALIDATION] Final stats - Processed: {job['processed']}, Verified: {job['verified']}, Partial: {job['partial']}, Not Found: {job['not_found']}, Errors: {job['errors']}")
            return

        job["status"] = "completed"
        job["current_status"] = "Validation complete"
        print(f"[AI VALIDATION] JOB COMPLETED: {job_id}")
        print(f"[AI VALIDATION] Final stats - Processed: {job['processed']}, Verified: {job['verified']}, Partial: {job['partial']}, Not Found: {job['not_found']}, Errors: {job['errors']}")
        
    except Exception as e:
        print(f"[AI VALIDATION] EXCEPTION in process_ai_validation_async: {e}")
        import traceback
        traceback.print_exc()
        job = ai_validation_jobs.get(job_id)
        if job:
            job["status"] = "error"
            job["current_status"] = f"Error: {str(e)}"
            print(f"[AI VALIDATION] Job marked as ERROR: {job_id}")


@app.route("/api/ai-validation/progress/<job_id>")
@login_required
def get_ai_validation_progress(job_id):
    """Get progress of an AI validation job."""
    try:
        job = ai_validation_jobs.get(job_id)
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        # Security check: ensure user owns this job
        if job["user"] != get_username():
            return jsonify({"error": "Unauthorized"}), 403
        
        # Use eligible_count for progress (eligible rows to process)
        # Use total for reference (total rows in worksheet)
        eligible_count = job.get("eligible_count", job["total"])
        
        response = {
            "status": job["status"],
            "total": job["total"],
            "eligible_count": eligible_count,
            "processed": job["processed"],
            "verified": job["verified"],
            "partial": job["partial"],
            "conflict": job["conflict"],
            "not_found": job["not_found"],
            "errors": job["errors"],
            "current_status": job["current_status"]
        }

        if job["status"] == "error":
            response["error"] = job.get("current_status", "Validation failed")
        
        # Include file_key when completed (so frontend knows file is registered in Files)
        if job["status"] == "completed" and job.get("file_key"):
            response["file_key"] = job["file_key"]
        
        return jsonify(response)
    
    except Exception as e:
        print(f"Error getting progress: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/ai-validation/download/<job_id>")
@login_required
def download_ai_validated_file(job_id):
    """Download the validated Excel file."""
    try:
        job = ai_validation_jobs.get(job_id)
        if not job:
            return jsonify({"error": "Job not found"}), 404

        if job["user"] != get_username():
            return jsonify({"error": "Unauthorized"}), 403

        if job["status"] != "completed":
            return jsonify({"error": "Job not completed"}), 400

        download_path = job.get("download_path")
        if not download_path or not os.path.exists(download_path):
            if job.get("file_key"):
                download_path = build_user_draft_path(job["user"], job["file_key"])
            if not download_path or not os.path.exists(download_path):
                if job.get("df") is not None:
                    download_path = build_user_draft_path(job["user"], job.get("file_key") or f"ai_validated_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
                    if not download_path:
                        return jsonify({"error": "Validated file could not be generated"}), 500
                    try:
                        safe_write_excel(job["df"], download_path, working_sheet_name=job.get("working_sheet_name", "Sheet1"), workbook_sheets=job.get("workbook_sheets"))
                    except Exception as exc:
                        return jsonify({"error": f"Validated file could not be generated: {str(exc)}"}), 500
                else:
                    return jsonify({"error": "Validated file not ready"}), 400

        if not os.path.exists(download_path):
            return jsonify({"error": "Validated file not found on disk"}), 500

        filename = os.path.basename(download_path)
        if filename.endswith("_draft.xlsx"):
            filename = filename.replace("_draft.xlsx", ".xlsx")
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        job["download_path"] = download_path
        return send_file(
            download_path,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error downloading file: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/ai-validation/revalidate/<key>", methods=["POST"])
@login_required
def revalidate_file(key):
    """Re-validate an existing file from the Files system."""
    print(f"[AI VALIDATION] revalidate_file called for key: {key}")
    
    try:
        user = get_username()
        if not is_safe_key(key):
            return jsonify({"error": "Invalid file key"}), 400
        
        # Verify file exists and belongs to user
        draft_path = get_user_draft_path(key)
        if not draft_path or not os.path.exists(draft_path):
            return jsonify({"error": "File not found"}), 404
        
        # Get validator from request
        data = request.get_json() or {}
        validator = data.get("validator", "").strip()
        if not validator:
            return jsonify({"error": "Validator required"}), 400
        
        print(f"[AI VALIDATION] Re-validating file: {key}, validator: {validator}")
        
        # Read the workbook
        try:
            with open(draft_path, "rb") as f:
                workbook_sheets = get_workbook_sheet_map(f)
            if not workbook_sheets:
                return jsonify({"error": "Could not read Excel file"}), 400
        except Exception as e:
            print(f"[AI VALIDATION] ERROR reading workbook: {e}")
            return jsonify({"error": f"Could not read Excel file: {str(e)}"}), 400
        
        # Detect working sheet (same logic as initial validation)
        working_sheet_name = detect_working_sheet_name(workbook_sheets) or next(iter(workbook_sheets), "Sheet1")
        df = workbook_sheets.get(working_sheet_name)
        
        if df is None or len(df) == 0:
            return jsonify({"error": "No data in workbook"}), 400
        
        # Generate job ID
        job_id = f"{user}_{int(time.time())}_{os.urandom(4).hex()}"
        
        # Initialize job (same as start_ai_validation)
        ai_validation_jobs[job_id] = {
            "status": "processing",
            "user": user,
            "validator": validator,
            "original_filename": key + ".xlsx",
            "df": df.copy(),
            "workbook_sheets": workbook_sheets,
            "working_sheet_name": working_sheet_name,
            "total": len(df),
            "eligible_rows": [],
            "eligible_count": 0,
            "processed": 0,
            "verified": 0,
            "partial": 0,
            "conflict": 0,
            "not_found": 0,
            "errors": 0,
            "current_status": "Initializing...",
            "created_at": datetime.now(),
            "download_path": None,
            "file_key": None
        }
        
        # Start background processing
        thread = threading.Thread(target=process_ai_validation_async, args=(job_id,), daemon=True)
        thread.start()
        
        print(f"[AI VALIDATION] Re-validation job started: {job_id}")
        return jsonify({"ok": True, "job_id": job_id})
        
    except Exception as e:
        print(f"[AI VALIDATION] ERROR in revalidate_file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
